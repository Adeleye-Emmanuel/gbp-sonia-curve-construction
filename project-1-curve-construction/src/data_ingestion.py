"""Load GBP SONIA OIS spot curve quotes from the Bank of England monthly archive.

Source: BoE "Monthly overnight index swap curve" archive
(https://www.bankofengland.co.uk/statistics/yield-curves), file
"OIS month end data_2025 to present.xlsx", sheet "4. spot curve ".

NOTE: these are BoE's Anderson-Sleath fitted spot rates, not raw traded par
swap quotes off a broker screen. We treat them as par-rate-equivalent inputs
to our bootstrap by design (already agreed) -- not genuine raw market quotes.
"""

import csv
from pathlib import Path

import openpyxl

RAW_XLSX = Path("data/raw/OIS month end data_2025 to present.xlsx")
SHEET_NAME = "4. spot curve "
PROCESSED_CSV = Path("data/processed/boe_ois_quotes.csv")


def extract_quotes(xlsx_path: Path = RAW_XLSX) -> list[tuple[float, float]]:
    """Extract whole-year (1-10Y) spot rates for the most recent date in the file.

    Returns (tenor_years, rate) pairs sorted ascending, rate as a decimal
    (e.g. 4.20 -> 0.042), in the exact format bootstrap_curve() expects.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    header = rows[3]
    last_row = rows[-1]

    quotes = []
    for tenor in range(1, 11):
        col_idx = header.index(tenor)
        rate_pct = last_row[col_idx]
        quotes.append((float(tenor), rate_pct / 100.0))
    return quotes


def save_quotes(quotes: list[tuple[float, float]], csv_path: Path = PROCESSED_CSV) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with open(csv_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["tenor_years", "rate"])
        writer.writerows(quotes)


if __name__ == "__main__":
    quotes = extract_quotes()
    save_quotes(quotes)
    print(quotes)
