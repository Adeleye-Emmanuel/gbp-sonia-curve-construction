"""Load GBP SONIA OIS daily spot curve history from the Bank of England.

Source: BoE "Daily overnight index swap curve" archive
(https://www.bankofengland.co.uk/statistics/yield-curves -> "Daily overnight
index swap curve: archive data"), unzipped into data/raw/oisddata_extracted/.
Files "OIS daily data_2016 to 2024.xlsx" and "OIS daily data_2025 to
present.xlsx", sheet "4. spot curve" (NOTE: unlike Project 1's monthly
file, this sheet name has no trailing space).

Needed for enough historical daily observations to run a meaningful PCA on
curve moves (Project 1's M3 only used the monthly archive, one curve per
month-end -- not enough observations for that).

Reconnaissance findings (inspected directly, not assumed):
- Header row (row 4) holds tenor years in 0.5Y increments, 0.5 through 25.0.
  Values are percent (e.g. 3.9096...), matching Project 1's convention of
  dividing by 100 to get a decimal rate.
- Column A holds the date; row 3 (0-indexed) is 'Maturity' label, row 4 is
  the tenor header, data starts row 5 (0-indexed row 4).
- The sheet has one row per *calendar* date, not just business days: UK
  bank holidays appear as a real row with every rate cell blank (None).
  Weekends are simply absent as rows. Excel's used-range also runs well
  past the last real row of data (trailing all-None rows) -- filtered out
  by requiring col A to be an actual date.
- The two files' data ranges are contiguous and non-overlapping (...2024
  ends 2024-12-31, ...present starts 2025-01-01), and share an identical
  tenor header.
"""

import csv
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

RAW_FILES = [
    Path("data/raw/oisddata_extracted/OIS daily data_2016 to 2024.xlsx"),
    Path("data/raw/oisddata_extracted/OIS daily data_2025 to present.xlsx"),
]
SHEET_NAME = "4. spot curve"
PROCESSED_CSV = Path("data/processed/boe_historical_curves.csv")
TENORS = list(range(1, 11))


def _extract_sheet(xlsx_path: Path) -> tuple[tuple, list[tuple]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    header = rows[3]
    data_rows = [r for r in rows[4:] if isinstance(r[0], datetime)]
    wb.close()
    return header, data_rows


def extract_historical_curves(raw_files: list[Path] = RAW_FILES, years: int = 2) -> list[dict]:
    """Extract whole-year (1-10Y) spot rates for the most recent `years` of
    business days available across the given archive files.

    Returns a list of dicts sorted ascending by date, one per business day,
    each {'date': date, '1Y': rate, ..., '10Y': rate}, rates as decimals.
    Bank-holiday rows (all rates blank) are dropped, not kept as gaps.
    """
    header = None
    all_rows = []
    for path in raw_files:
        sheet_header, data_rows = _extract_sheet(path)
        if header is None:
            header = sheet_header
        elif header != sheet_header:
            raise ValueError(f"tenor header mismatch in {path}")
        all_rows.extend(data_rows)

    tenor_col = {tenor: header.index(float(tenor)) for tenor in TENORS}

    max_date = max(row[0] for row in all_rows)
    cutoff = max_date - timedelta(days=365 * years)

    quotes = []
    for row in all_rows:
        row_date = row[0]
        if row_date <= cutoff:
            continue
        rates = [row[tenor_col[tenor]] for tenor in TENORS]
        if any(rate is None for rate in rates):
            continue
        quotes.append(
            {"date": row_date.date(), **{f"{tenor}Y": rate / 100.0 for tenor, rate in zip(TENORS, rates)}}
        )

    quotes.sort(key=lambda q: q["date"])
    return quotes


def save_historical_curves(quotes: list[dict], csv_path: Path = PROCESSED_CSV) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["date"] + [f"{tenor}Y" for tenor in TENORS]
    with open(csv_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in quotes:
            writer.writerow(row)


if __name__ == "__main__":
    quotes = extract_historical_curves()
    save_historical_curves(quotes)
    print(f"{len(quotes)} rows, {quotes[0]['date']} to {quotes[-1]['date']}")
